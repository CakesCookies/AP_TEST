import pandas as pd
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Side, Border
from GlobalVar import *


def Only_Read(File_Name):
    Txt_program = open(File_Name, 'r')
    Output = Txt_program.read()
    Txt_program.close
    return Output


def DataFrame_Create(Row_List, Column_List):
    Output_df = pd.DataFrame(columns=Column_List, index=Row_List, dtype='str')
    Output_df[:] = ''
    return Output_df


def Draw_Info_Sheet(Input_df, WorkBook, Sheet_Name):
    ColorEnable = False
    if ColorEnable == True:
        SummaryColor_STR = 'EFBD1F'
    STR_Size = 12
    WorkBook.create_sheet(Sheet_Name)
    WorkSheet = WorkBook[Sheet_Name]
    if ColorEnable == True:
        WorkSheet.sheet_properties.tabColor = SummaryColor_STR
    TitleSet_Column = 1
    # Title
    for Index_Excel, Index_STR in enumerate(Input_df.index):
        Cell_col = TitleSet_Column
        Cell_row = Index_Excel + 1
        WorkSheet.cell(column=Cell_col, row=Cell_row).value = Index_STR
        if ColorEnable == True:
            if Index_STR in Summary_Item_Get():
                WorkSheet.cell(column=Cell_col, row=Cell_row).fill = PatternFill('solid', fgColor=SummaryColor_STR)
            else:
                WorkSheet.cell(column=Cell_col, row=Cell_row).fill = PatternFill('solid', fgColor='C2E6FF')
        WorkSheet.cell(column=Cell_col, row=Cell_row).font = Font(color='000000', size=STR_Size)
        WorkSheet.cell(column=Cell_col, row=Cell_row).alignment = Alignment(horizontal='left',
                                                                                            vertical='center')
        WorkSheet.cell(column=Cell_col, row=Cell_row).border = Frame_Style_Thin_Get()
    # Value
    TitleSet_Column = TitleSet_Column + 1
    for Index_Excel, Index_STR in enumerate(Input_df.index):
        Cell_col = TitleSet_Column
        Cell_row = Index_Excel + 1
        WorkSheet.cell(column=Cell_col, row=Cell_row).value = Input_df.at[Index_STR,'Value']
        WorkSheet.cell(column=Cell_col, row=Cell_row).font = Font(color='000000', size=STR_Size)
        WorkSheet.cell(column=Cell_col, row=Cell_row).alignment = Alignment(horizontal='center',
                                                                                            vertical='center')
        WorkSheet.cell(column=Cell_col, row=Cell_row).border = Frame_Style_Thin_Get()
    AutoColumns(WorkSheet, STR_Size)
    return WorkBook


def Draw_Result_Sheet(Input_df, WorkBook, Sheet_Name):
    ColorEnable = False
    if ColorEnable == True:
        SpaceColor_STR = 'C2E6FF'
        FailColor_STR='FF0000'
    STR_Size = 12
    WorkBook.create_sheet(Sheet_Name)
    WorkSheet = WorkBook[Sheet_Name]
    TitleSet_Row = 1
    TitleSet_Column = 1
    # Title -> Column
    if ColorEnable == True:
        WorkSheet.cell(column=TitleSet_Row, row=TitleSet_Column).fill = PatternFill('solid', fgColor='FFE6FF')
        if len(Input_df[Input_df['Result']=='Fail']) >= 1:
            WorkSheet.sheet_properties.tabColor = FailColor_STR
    for Column_Excel, Column_STR in enumerate(Input_df):
        Cell_col = Column_Excel + TitleSet_Column + 1
        Cell_row = TitleSet_Row
        WorkSheet.cell(column=Cell_col, row=Cell_row).value = Column_STR
        if ColorEnable == True:
            WorkSheet.cell(column=Cell_col, row=Cell_row).fill = PatternFill('solid', fgColor='FFE6FF')
        WorkSheet.cell(column=Cell_col, row=Cell_row).font = Font(color='000000', size=STR_Size)
        WorkSheet.cell(column=Cell_col, row=Cell_row).alignment = Alignment(horizontal='center',
                                                                                            vertical='center')
        WorkSheet.cell(column=Cell_col, row=Cell_row).border = Frame_Style_Thin_Get()
    # Title -> Row
    for Index_Excel, Index_STR in enumerate(Input_df.index):
        Cell_col = TitleSet_Column
        Cell_row = TitleSet_Row + Index_Excel + 1
        WorkSheet.cell(column=Cell_col, row=Cell_row).value = Index_STR
        if ColorEnable == True:
            if (Cell_row % 2) == 1:
                WorkSheet.cell(column=Cell_col, row=Cell_row).fill = PatternFill('solid', fgColor=SpaceColor_STR)
        WorkSheet.cell(column=Cell_col, row=Cell_row).font = Font(color='000000', size=STR_Size)
        WorkSheet.cell(column=Cell_col, row=Cell_row).alignment = Alignment(horizontal='left',  vertical='center')
        WorkSheet.cell(column=Cell_col, row=Cell_row).border = Frame_Style_Thin_Get()
    # Text
    for Index_Excel, Index_STR in enumerate(Input_df.index):
        for Column_Excel, Column_STR in enumerate(Input_df):
            Cell_col = TitleSet_Column + Column_Excel + 1
            Cell_row = TitleSet_Row + Index_Excel + 1
            WorkSheet.cell(column=Cell_col, row=Cell_row).value = Input_df.at[Index_STR, Column_STR]
            if ColorEnable == True:
                if Column_STR == 'Result' and Input_df.at[Index_STR, Column_STR] == 'Fail':
                    WorkSheet.cell(column=Cell_col, row=Cell_row).fill = PatternFill('solid', fgColor=FailColor_STR)
                elif (Cell_row % 2) == 1:
                    WorkSheet.cell(column=Cell_col, row=Cell_row).fill = PatternFill('solid',
                                                                                                     fgColor=SpaceColor_STR)
            WorkSheet.cell(column=Cell_col, row=Cell_row).font = Font(color='000000', size=STR_Size)
            WorkSheet.cell(column=Cell_col, row=Cell_row).alignment = Alignment(horizontal='center',
                                                                                                vertical='center')
            WorkSheet.cell(column=Cell_col, row=Cell_row).border = Frame_Style_Thin_Get()
    AutoColumns(WorkSheet, STR_Size)
    return WorkBook


def Frame_Style_Thin_Get():
    B_Border = Border(  left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000'))
    return B_Border


def AutoColumns(WorkSheet, STR_Size):
    Indent = 2
    for Col_i in WorkSheet.columns:
        MaxSize = 0
        Col_STR = Col_i[0].column
        for cell_ij in Col_i:
            if cell_ij.value:
                MaxSize = max(MaxSize, len(str(cell_ij.value)))
        WorkSheet.column_dimensions[Col_STR].width = int ( (MaxSize + Indent) * STR_Size * 0.12 )
    return WorkSheet


def List_To_DataFrame(Data_List):
    DataFrame_Panel = {}
    Info_df = DataFrame_Create(['Device Brand', 'Device Name', 'Android Version', 'AP Type', ], ['Value'])
    CurrentLoop = '1'
    for Dara_i in Data_List:
        if 'Device Brand' in Dara_i:
            Info_df.at['Device Brand', 'Value'] = Dara_i.split(':')[1].replace(' ', '', 1)
        elif 'Device Name' in Dara_i:
            Info_df.at['Device Name', 'Value'] = Dara_i.split(':')[1].replace(' ', '', 1)
        elif 'Android Version' in Dara_i:
            Info_df.at['Android Version', 'Value'] = Dara_i.split(':')[1].replace(' ', '', 1)
        elif 'AP Type' in Dara_i:
            Info_df.at['AP Type', 'Value'] = Dara_i.split(':')[1].replace(' ', '', 1)
        elif 'Test loop' in Dara_i:
            CurrentLoop = Dara_i.split(':')[1].replace(' ', '', 1)

        if CurrentLoop not in DataFrame_Panel:
            DataFrame_Panel[CurrentLoop] = DataFrame_Create(None,
                                                            ['Time', 'Band', 'Security', 'Standard', 'MHz', 'Channels',
                                                             'Test Case', 'Action', 'Result'])
        if ('Pass' in Dara_i) or ('Fail' in Dara_i):
            Data_Spilt = Dara_i.split('_')
            Time_Time = datetime.datetime.strptime(Data_Spilt[0].split(' ')[0] + ' ' + Data_Spilt[0].split(' ')[1],
                                                   '%Y/%m/%d %H:%M:%S')
            AddIndex = str(len(DataFrame_Panel[CurrentLoop].index) + 1)
            DataFrame_Panel[CurrentLoop].loc[AddIndex] = [Time_Time, Data_Spilt[1], Data_Spilt[2], Data_Spilt[3],
                                                          Data_Spilt[4], Data_Spilt[5], Data_Spilt[6], Data_Spilt[7], Data_Spilt[8][1:]]
    return Info_df, DataFrame_Panel

def CalculateSummary(Info_df, DataFrame_Panel):
    All_df = DataFrame_Panel[ list( DataFrame_Panel.keys() )[0] ][:0]
    for Loop_i in DataFrame_Panel:
        All_df = pd.concat([All_df, DataFrame_Panel[Loop_i]])
    Summary_List = Summary_Item_Get()
    Info_df.loc[Summary_List[0]] = min(All_df['Time'])
    Info_df.loc[Summary_List[1]] = max(All_df['Time'])
    PassItem = len(All_df[All_df['Result']=='Pass'])
    FailItem = len(All_df[All_df['Result']=='Fail'])
    Info_df.loc[Summary_List[2]] = PassItem
    Info_df.loc[Summary_List[3]] = FailItem
    Info_df.loc[Summary_List[4]] = round(FailItem/(PassItem + FailItem) , 2)
    return Info_df

def Summary_Item_Get():
    Add_Item = ['Start Time', 'End Time', 'Pass Item', 'Fail Item', 'Fail Rate']
    return Add_Item

def DataFrame_To_Excel(Info_df, DataFrame_Panel, ExcelName):
    WorkBook = Workbook()
    WorkBook = Draw_Info_Sheet(Info_df, WorkBook, 'Result')
    for Data_i in DataFrame_Panel:
        WorkBook = Draw_Result_Sheet(DataFrame_Panel[Data_i], WorkBook, Data_i)
    WorkBook.remove_sheet(WorkBook['Sheet'])
    WorkBook.save(ExcelName)
    return


def TXTFile_To_ExcelFile(File_Name):
    Output_ExcelFile = File_Name.replace('.txt', '.xlsx')
    DataAll = Only_Read(File_Name)
    Data_List = DataAll.split('\n')
    Surroundings_df, Data_Panel = List_To_DataFrame(Data_List)
    Surroundings_df = CalculateSummary(Surroundings_df, Data_Panel)
    DataFrame_To_Excel(Surroundings_df, Data_Panel, Output_ExcelFile)
    return


if __name__ == '__main__':
    File_Name = 'Test Report_20220809-190000.txt'
    TXTFile_To_ExcelFile(File_Name)
